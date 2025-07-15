import meals
import os
from openpyxl import load_workbook
from flask import Flask, render_template



def read_from_excel():
            wb = load_workbook('CalorieTracker.xlsx')

            ws = wb['Macros']
            tracked_meals = []

            headers = [cell.value for cell in ws[1]]

            # print(headers)
            #Appending the Meals to the Meals dictionary
            for row in ws.iter_rows(min_row=2, values_only=True):
                meal = {headers[i]: row[i] for i in range(len(headers))}
                tracked_meals.append(meal)
            # print(meals)
            return tracked_meals


#Data structures method

tracked_meals = read_from_excel()
tracked_meals = [meal for meal in tracked_meals if meal['Food']  is not None]

# unique_foods = set(meal['Protein'] for meal in tracked_meals if meal['Protein'] is not None)
# print(unique_foods)

# sorted_meals = sorted(tracked_meals, key=lambda meal: int(meal['Protein']), reverse=True)


# protein_values = [meal['Protein'] for meal in tracked_meals if meal['Protein'] is not None]
# print(protein_values)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "..", "templates")

app = Flask(__name__, template_folder=TEMPLATE_DIR )

print("template folder = ")
print( app.template_folder )
print("🧭 Flask is looking for templates in:", app.template_folder)

if os.path.exists(app.template_folder ):
    print("📂 Found templates folder. Files inside:")
    for file in os.listdir(app.template_folder):
        print(" -", file)
else:
    print("❌ templates folder NOT FOUND at this location.")


@app.route('/')

def home():
      return render_template("input.html")


def dashboard():
        # print("Meals to be tracked ")
        
        print(tracked_meals)

        return render_template('index.html', meals = tracked_meals)


if __name__ == '__main__':
         app.run(debug=True)